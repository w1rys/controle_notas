import os
import pandas as pd
from utils import log, log_erro
from openpyxl import load_workbook


# ----------------------------------------------------------
# Carregar Excel
# ----------------------------------------------------------
def carregar_excel(nome_excel):
    """Carrega ou cria o Excel com a aba 'Compras'."""

    colunas_base = [
        "codigo",
        "nome_produto",
        "quantidade_total_comprada",
        "ultimo_preco",
        "penultimo_preco",
        "ultima_compra",
        "chave_nota"
    ]

    if not os.path.exists(nome_excel):
        df = pd.DataFrame(columns=colunas_base)
        df.to_excel(nome_excel, sheet_name="Compras", index=False)
        log("[INFO] Arquivo Excel criado com aba 'Compras'.")
        return df

    try:
        df = pd.read_excel(nome_excel, sheet_name="Compras", parse_dates=["ultima_compra"])

        # adicionar colunas ausentes
        for col in colunas_base:
            if col not in df.columns:
                df[col] = None

        return df

    except ValueError:
        # Excel existe mas não tem aba Compras
        df = pd.DataFrame(columns=colunas_base)
        with pd.ExcelWriter(nome_excel, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="Compras", index=False)
        return df


# ----------------------------------------------------------
# Garantir datas sem timezone + salvar aba Compras
# ----------------------------------------------------------
def salvar_excel(df_produtos, nome_excel):
    try:
        # Garantir datetime sem timezone
        if "ultima_compra" in df_produtos.columns:
            df_produtos["ultima_compra"] = pd.to_datetime(df_produtos["ultima_compra"], errors="coerce")

            try:
                if df_produtos["ultima_compra"].dt.tz is not None:
                    df_produtos["ultima_compra"] = df_produtos["ultima_compra"].dt.tz_convert(None)
            except:
                df_produtos["ultima_compra"] = df_produtos["ultima_compra"].apply(
                    lambda t: t.replace(tzinfo=None) if hasattr(t, "tzinfo") and t.tzinfo else t
                )

        # Criar se não existir arquivo
        if not os.path.exists(nome_excel):
            df_produtos.to_excel(nome_excel, sheet_name="Compras", index=False)
            log("[INFO] Excel criado com aba 'Compras'.")
            return

        # Atualizar aba Compras
        with pd.ExcelWriter(
            nome_excel,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            df_produtos.to_excel(writer, sheet_name="Compras", index=False)

        log("[INFO] Aba 'Compras' atualizada.")

    except Exception as e:
        log_erro(f"Erro ao salvar o Excel: {e}")


# ----------------------------------------------------------
# Criar/Atualizar aba "Produtos"
# ----------------------------------------------------------
def atualizar_aba_produtos(df_produtos, nome_excel="produtos.xlsx"):
    """
    ABA "Produtos":
    - codigo
    - nome_produto
    - ultima_compra (datetime sem hora no Excel)
    - ultimo_preco
    - penultimo_preco
    - preco_venda
    """

    # Garantir datetime sem timezone e sem hora
    if "ultima_compra" in df_produtos.columns:
        df_produtos["ultima_compra"] = (
            pd.to_datetime(df_produtos["ultima_compra"], errors="coerce")
            .dt.tz_localize(None)
            .dt.normalize()     # <<< Remove hora
        )

    # Ordem das colunas
    df_base = df_produtos[[
        "codigo",
        "nome_produto",
        "ultima_compra",
        "ultimo_preco",
        "penultimo_preco"
    ]].copy()

    # Preencher penúltimo preço quando necessário
    df_base["penultimo_preco"] = df_base.apply(
        lambda row: row["ultimo_preco"]
        if pd.isna(row["penultimo_preco"]) else row["penultimo_preco"],
        axis=1
    )

    # Ordenar alfabeticamente
    df_base = df_base.sort_values(by="nome_produto")

    aba = "Produtos"

    try:
        book = load_workbook(nome_excel)

        # Preservar preco_venda
        if aba in book.sheetnames:
            antigo = pd.read_excel(nome_excel, sheet_name=aba)
            if "preco_venda" in antigo.columns:
                df_base = df_base.merge(
                    antigo[["codigo", "preco_venda"]],
                    on="codigo",
                    how="left"
                )
            else:
                df_base["preco_venda"] = None
        else:
            df_base["preco_venda"] = None

        # Escrever a aba
        with pd.ExcelWriter(
            nome_excel,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            df_base.to_excel(writer, sheet_name=aba, index=False)

        # ---------------------------
        # FORMATAÇÃO DA DATA NO EXCEL
        # ---------------------------
        book = load_workbook(nome_excel)
        ws = book[aba]

        COL_DATA = 3  # código=1, nome=2, data=3

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=COL_DATA)
            cell.number_format = "DD/MM/YYYY"   # <<< sem hora

        book.save(nome_excel)

    except FileNotFoundError:
        df_base["preco_venda"] = None

        with pd.ExcelWriter(
            nome_excel,
            engine="openpyxl",
            mode="w"
        ) as writer:
            df_base.to_excel(writer, sheet_name=aba, index=False)

        # Format after creating
        book = load_workbook(nome_excel)
        ws = book[aba]

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            cell.number_format = "DD/MM/YYYY"

        book.save(nome_excel)
# ----------------------------------------------------------
# Atualizar Excel após leitura da nota
# ----------------------------------------------------------
def atualizar_excel_compras(novos_produtos, chave_nota, nome_excel="produtos.xlsx"):
    df = carregar_excel(nome_excel)

    # Evitar duplicidade
    if chave_nota in df["chave_nota"].astype(str).values:
        log(f"[INFO] Nota {chave_nota} já registrada – ignorando.")
        return

    for item in novos_produtos:
        codigo = item["codigo"]
        qtd_nova = item["quantidade"]
        preco_novo = item["preco_unitario"]
        data_nova = pd.to_datetime(item["data_compra"]).tz_localize(None)

        # Verificar se produto já existe
        existente = df[df["codigo"] == codigo]

        if not existente.empty:
            idx = existente.index[0]

            # Quantidade acumulada
            qtd_antiga = df.loc[idx, "quantidade_total_comprada"]
            df.loc[idx, "quantidade_total_comprada"] = qtd_antiga + qtd_nova

            # Data registrada atualmente
            data_atual = df.loc[idx, "ultima_compra"]
            data_atual = pd.to_datetime(data_atual) if not pd.isna(data_atual) else None

            # ----------------------------------------------------------
            # ATUALIZAÇÃO BASEADA NA DATA DA NOTA (CORRETO!)
            # ----------------------------------------------------------

            # Se a nota nova for mais RECENTE
            if data_atual is None or data_nova > data_atual:

                # penúltimo preço recebe o último
                ultimo_antigo = df.loc[idx, "ultimo_preco"]
                df.loc[idx, "penultimo_preco"] = ultimo_antigo

                # último preço recebe o novo valor
                df.loc[idx, "ultimo_preco"] = preco_novo

                # data da última compra atualizada
                df.loc[idx, "ultima_compra"] = data_nova

            else:
                # Nota antiga → não altera preços
                log(f"[INFO] Nota antiga detectada para {codigo}: preço não atualizado.")

        else:
            # Produto NOVO → cadastra tudo
            df.loc[len(df)] = {
                "codigo": codigo,
                "nome_produto": item["nome_produto"],
                "quantidade_total_comprada": qtd_nova,
                "ultimo_preco": preco_novo,
                "penultimo_preco": preco_novo,
                "ultima_compra": data_nova,
                "chave_nota": chave_nota
            }

    # ordenar alfabeticamente
    df = df.sort_values(by="nome_produto", ascending=True)

    # salvar Excel
    salvar_excel(df, nome_excel)

    # atualizar aba Produtos com datetime real
    atualizar_aba_produtos(df, nome_excel)
